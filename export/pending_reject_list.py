'''
Purpose: This module handles the export of pending and rejected incident data from MongoDB to formatted Excel reports
Created Date: 2025-01-18
Created By: Aruna Jayaweera (ajayaweerau@gmail.com)
Last Modified Date: 2024-02-20
Modified By: Aruna Jayaweera (ajayaweerau@gmail.com)
Version: Python 3.12
Dependencies:
    - pymongo (for MongoDB connectivity)
    - openpyxl (for Excel file operations)
    - python-dotenv (for environment variables)
    - logging (for execution tracking)
Related Files:
    - task_handler.py (initiates the export process)
    - config_loader.py (provides export path configuration)
    - style_loader.py (handles Excel styling)
    - connectionMongo.py (database connection handler)

Program Description:
1. Core Functionality:
    - excel_pending_reject_incident(): Main export function that:
        a. Validates input parameters (DRC commission rules, date range)
        b. Constructs MongoDB query for pending/rejected incidents
        c. Executes query against Incident_log collection
        d. Generates formatted Excel report
    - create_pending_reject_incident_table(): Handles Excel sheet creation with:
        a. Professional formatting and styling
        b. Dynamic column sizing
        c. Filter headers display
        d. Empty dataset handling

2. Data Flow:
    - Receives filter parameters from calling function
    - Fetches data from Incident_log collection with "Incident_Status" in ["Incident Pending", "Incident Reject"]
    - Transforms MongoDB documents to Excel rows
    - Applies consistent styling using STYLES configuration
    - Saves report to configured export directory

3. Key Features:
    - Parameter Validation:
        - Validates DRC commission rules as non-empty list
        - Date format enforcement (YYYY-MM-DD)
        - Date range validation (to_date cannot be earlier than from_date)
    - Error Handling:
        - Comprehensive validation errors
        - Database operation failures
        - File system permissions
    - Reporting:
        - Automatic filename generation with timestamp (pending_reject_incidents_[timestamp].xlsx)
        - Empty dataset handling with headers
        - Console and log feedback

4. Configuration:
    - Export path determined by ConfigLoaderSingleton
    - Styles managed through style_loader.py
    - Column headers defined in PENDING_REJECT_INCIDENT_HEADERS constant:
        * Incident_Id
        * Incident_Status
        * Account_Num
        * Filtered_Reason
        * Rejected_Dtm
        * Source_Type

5. Integration Points:
    - Called by task handlers for pending/rejected incident reporting
    - Uses MongoDBConnectionSingleton for database access
    - Leverages application-wide logging

Technical Specifications:
    - Input Parameters:
        - drc_commission_rules: List of strings (valid commission rules)
        - from_date/to_date: String (YYYY-MM-DD format)
    - Output:
        - Excel file with standardized naming convention
        - Returns boolean success status
    - Collections Accessed:
        - Incident_log (primary data source)
        - Filters for status in ["Incident Pending", "Incident Reject"]
    - Special Data Handling:
        - Converts ObjectId to string for Incident_Id
        - Formats datetime objects for Rejected_Dtm
        - Maintains headers even with empty result sets
'''

from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from utils.connectionMongo import MongoDBConnectionSingleton
from logging import getLogger
from tasks.config_loader import ConfigLoaderSingleton

logger = getLogger('appLogger')

PENDING_REJECT_INCIDENT_HEADERS = [
    "Incident_Id", "Incident_Status", "Account_Num", "Filtered_Reason",
    "Rejected_Dtm", "Source_Type"
]

def excel_pending_reject_incident(drc_commission_rules, from_date, to_date):
    """Fetch and export pending reject incidents based on validated parameters"""
    
    try:
            #Get export directory from config
            export_dir = ConfigLoaderSingleton().get_export_path()
            export_dir.mkdir(parents=True, exist_ok=True)

            db = MongoDBConnectionSingleton().get_database()
            incident_log_collection = db["Incident"]

            pending_reject_query = {"Incident_Status": {"$in": [ "Incident Reject"]}}

            # Check drc_commission_rules
            if drc_commission_rules is not None:
                if isinstance(drc_commission_rules, list) and drc_commission_rules:
                    pending_reject_query["Filtered_Reason"] = {"$in": drc_commission_rules}
                else:
                    raise ValueError("drc_commission_rules must be a non-empty list of valid commission rules")

            # Check date range
            if from_date is not None and to_date is not None:
                try:
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                    
                    if to_dt < from_dt:
                        raise ValueError("to_date cannot be earlier than from_date")
                    
                    pending_reject_query["Rejected_Dtm"] = {"$gte": from_dt, "$lte": to_dt}

                except ValueError as ve:
                    if str(ve).startswith("to_date"):
                        raise
                    raise ValueError(f"Invalid date format. Use 'YYYY-MM-DD'. Error: {str(ve)}")

            # Log and execute query
            logger.info(f"Executing query: {pending_reject_query}")
            incidents = list(incident_log_collection.find(pending_reject_query))
            logger.info(f"Found {len(incidents)} matching incidents")

            # Export to Excel even if no incidents are found
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
            filename = f"pending_reject_incidents_{timestamp}.xlsx"
            filepath = export_dir / filename

            wb = Workbook()
            wb.remove(wb.active)

            if not create_pending_reject_incident_table(wb, incidents, {
                "drc_commission_rules": drc_commission_rules,
                "date_range": (from_dt if from_date is not None else None, to_dt if to_date is not None else None)
            }):
                raise Exception("Failed to create pending reject incident sheet")

            wb.save(filepath)

             # Write export record to Download collection
            try:
                download_collection = db["download"]
                export_record = {
                    "File_Name": filename,
                    "File_Path": str(filepath),
                    "Export_Timestamp": datetime.now(),
                    "Exported_Record_Count": len(incidents),
                    "Applied_Filters": {
                        "DRC_Commision_Rule": drc_commission_rules,
                        "From_Date": from_date,
                        "To_Date": to_date
                    }
                }
                download_collection.insert_one(export_record)
                logger.info("Export details written to Download collection.")
            except Exception as e:
                logger.error(f"Failed to insert download record: {str(e)}", exc_info=True)


            if not incidents:
                print(f"No pending reject incidents found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(incidents)} records to: {filepath}")
            return True

    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}")
        print(f"Error: {str(ve)}")
        return False
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False
        
def create_pending_reject_incident_table(wb, data, filters=None):
    """Create formatted Excel sheet with pending reject incident data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="PENDING REJECT INCIDENT REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(PENDING_REJECT_INCIDENT_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="PENDING REJECT INCIDENT REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            if filters.get('drc_commission_rules'):
                ws.cell(row=row_idx, column=2, value="DRC Commission Rules:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=", ".join(filters['drc_commission_rules'])).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('date_range') and any(filters['date_range']):
                start, end = filters['date_range']
                ws.cell(row=row_idx, column=2, value="Date Range:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                date_str = f"{start.strftime('%Y-%m-%d') if start else 'Beginning'} to {end.strftime('%Y-%m-%d') if end else 'Now'}"
                ws.cell(row=row_idx, column=3, value=date_str).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(PENDING_REJECT_INCIDENT_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(PENDING_REJECT_INCIDENT_HEADERS, 1):
                    value = record.get(header, "")
                    if header == "Incident_Id" and isinstance(value, ObjectId):
                        value = str(value)
                    if header == "Rejected_Dtm" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(PENDING_REJECT_INCIDENT_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(PENDING_REJECT_INCIDENT_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating sheet: {str(e)}", exc_info=True)
        return False
