'''
Purpose: This module handles the export of case data from MongoDB to formatted Excel reports
Created Date: 2025-01-18
Created By: Aruna Jayaweera (ajayaweerau@gmail.com)
Last Modified Date: 2024-02-20
Modified By: Aruna Jayaweera (ajayaweerau@gmail.com)
Version: Python 3.12
Dependencies:
    - pymongo (for MongoDB connectivity)
    - openpyxl (for Excel file operations)
    - python-dotenv (for environment variables)
    - logging (for execution tracking)
Related Files:
    - task_handler.py (initiates the export process)
    - config_loader.py (provides export path configuration)
    - style_loader.py (handles Excel styling)
    - connectionMongo.py (database connection handler)

Program Description:
1. Core Functionality:
    - excel_case_detail(): Main export function that:
        a. Validates input parameters (case status, date range)
        b. Constructs MongoDB query for case data
        c. Executes query against Case_log collection
        d. Generates formatted Excel report
    - create_case_table(): Handles Excel sheet creation with:
        a. Professional formatting and styling
        b. Dynamic column sizing
        c. Filter headers display
        d. Empty dataset handling

2. Data Flow:
    - Receives filter parameters from calling function
    - Fetches data from Case_log collection
    - Transforms MongoDB documents to Excel rows
    - Applies consistent styling using STYLES configuration
    - Saves report to configured export directory

3. Key Features:
    - Parameter Validation:
        - Valid case statuses: "Pending FMB", "In progress", "Closed"
        - Date format enforcement (YYYY-MM-DD)
        - Date range validation (date_to cannot be earlier than date_from)
    - Special Data Handling:
        - Formats date fields (mm/dd/YYYY)
        - Combines validity period dates into single string
    - Error Handling:
        - Comprehensive validation errors
        - Database operation failures
        - File system permissions
    - Reporting:
        - Automatic filename generation with timestamp (cases_details_[timestamp].xlsx)
        - Empty dataset handling with headers
        - Console and log feedback

4. Configuration:
    - Export path determined by ConfigLoaderSingleton
    - Styles managed through style_loader.py
    - Column headers defined in CASE_HEADERS constant:
        * Case ID
        * Status
        * Request status
        * Validity Period
        * DRC
        * Request Details
        * Letter issued on
        * Approved on
        * Approved by
        * Remark

5. Integration Points:
    - Called by task handlers for case reporting
    - Uses MongoDBConnectionSingleton for database access
    - Leverages application-wide logging

Technical Specifications:
    - Input Parameters:
        - case_current_status: String (predefined values)
        - date_from/date_to: String (YYYY-MM-DD format)
    - Output:
        - Excel file with standardized naming convention
        - Returns boolean success status
    - Collections Accessed:
        - Case_log (primary data source)
    - Query Logic:
        - Checks both "Approved on" and "Letter issued on" dates
        - Uses exact matching for status values
'''

from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from utils.connectionMongo import MongoDBConnectionSingleton
from logging import getLogger
from utils.config_loader import ConfigLoaderSingleton

logger = getLogger('appLogger')

CASE_HEADERS = [
    "Case ID", "Status", "Request status", "Validity Period", "DRC",
    "Request Details", "Letter issued on", "Approved on", "Approved by", "Remark"
]

def excel_case_detail(case_current_status, date_from, date_to):
    """Fetch and export case data based on validated parameters"""
    
    try:   
            # Get export directory from config
            export_dir = ConfigLoaderSingleton().get_export_path()
            export_dir.mkdir(parents=True, exist_ok=True)

            db = MongoDBConnectionSingleton().get_database()
            collection = db["Case_log"]
            query = {} 

            # Check case_current_status parameter
            if case_current_status is not None:
                if case_current_status == "Pending FMB":
                     query["Status"] == {"$regex": f"^{case_current_status}$"}
                elif case_current_status == "In progress":
                     query["Status"] == case_current_status
                elif case_current_status == "Closed":
                     query["Status"] == case_current_status     
                else:
                    raise ValueError(f"Invalid case_current_status '{case_current_status}'. Must be 'Pending FMB', 'In progress', 'Closed' ")
                

            # Check date range
            if date_from is not None and date_to is not None:
                try:
                    # Check if dates are in correct YYYY-MM-DD format
                    from_dt = datetime.strptime(date_from, '%Y-%m-%d')
                    to_dt = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                    
                    # Validate date range
                    if to_dt < from_dt:
                        raise ValueError("date_to cannot be earlier than date_from")
                    
                    # Construct query for date range (checking both Approved on and Letter issued on)
                    query["$or"] = [
                        {"Approved on": {"$gte": from_dt, "$lte": to_dt}},
                        {"Letter issued on": {"$gte": from_dt, "$lte": to_dt}}
                    ]

                except ValueError as ve:
                    if str(ve).startswith("date_to"):
                        raise
                    raise ValueError(f"Invalid date format. Use 'YYYY-MM-DD'. Error: {str(ve)}")

            # Log and execute query
            logger.info(f"Executing query: {query}")
            cases = list(collection.find(query))
            logger.info(f"Found {len(cases)} matching cases")

            # Export to Excel even if no cases are found
            output_dir = "exports"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"cases_details_{timestamp}.xlsx"
            filepath = export_dir / filename

            wb = Workbook()
            wb.remove(wb.active)

            if not create_case_table(wb, cases, {
                "status": case_current_status,
                "date_range": (from_dt if date_from is not None else None, to_dt if date_to is not None else None)
            }):
                raise Exception("Failed to create case sheet")

            wb.save(filepath)
            if not cases:
                print(f"No cases found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(cases)} records to: {filepath}")
            return True

    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}")
        print(f"Error: {str(ve)}")
        return False
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False
        

def create_case_table(wb, data, filters=None):
    """Create formatted Excel sheet with filtered case data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="CASE REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(CASE_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="CASE REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            if filters.get('status'):
                ws.cell(row=row_idx, column=2, value="Status:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['status']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('date_range') and any(filters['date_range']):
                start, end = filters['date_range']
                ws.cell(row=row_idx, column=2, value="Date Range:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                date_str = f"{start.strftime('%Y-%m-%d') if start else 'Beginning'} to {end.strftime('%Y-%m-%d') if end else 'Now'}"
                ws.cell(row=row_idx, column=3, value=date_str).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(CASE_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(CASE_HEADERS, 1):
                    value = record.get(header, "")
                    # Handle date fields
                    if header in ["Letter issued on", "Approved on"] and isinstance(value, datetime):
                        value = value.strftime('%m/%d/%Y')
                    # Handle Validity Period if it's stored as separate dates
                    if header == "Validity Period":
                        start_date = record.get("Validity Period Start", "")
                        end_date = record.get("Validity Period End", "")
                        if start_date and end_date:
                            value = f"{start_date.strftime('%m/%d/%Y')} - {end_date.strftime('%m/%d/%Y')}"
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(CASE_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(CASE_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating sheet: {str(e)}", exc_info=True)
        return False