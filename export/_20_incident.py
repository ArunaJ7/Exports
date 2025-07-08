'''
Purpose: This module handles the export of incident data from MongoDB to formatted Excel reports
Created Date: 2025-03-20
Created By: Aruna Jayaweera (ajayaweerau@gmail.com)
Last Modified Date: 2025-05-20
Modified By: Aruna Jayaweera (ajayaweerau@gmail.com)
Version: Python 3.12
Dependencies:
    - pymongo (for MongoDB connectivity)
    - openpyxl (for Excel file operations)
    - python-dotenv (for environment variables)
    - logging (for execution tracking)
Related Files:
    - task_handler.py (initiates the export process)
    - config_loader.py (provides export path configuration)
    - style_loader.py (handles Excel styling)
    - connectionMongo.py (database connection handler)

Program Description:
1. Core Functionality:
    - excel_incident_detail(): Main export function that:
        a. Validates input parameters (action_type, status, date range)
        b. Constructs MongoDB query based on filters
        c. Executes query and processes results
        d. Generates formatted Excel report
    - create_incident_table(): Handles Excel sheet creation with:
        a. Professional formatting and styling
        b. Dynamic column sizing
        c. Filter headers display
        d. Empty dataset handling

2. Data Flow:
    - Receives filter parameters from TaskHandler
    - Fetches data from Incident_log collection
    - Transforms MongoDB documents to Excel rows
    - Applies consistent styling using STYLES configuration
    - Saves report to configured export directory

3. Key Features:
    - Parameter Validation:
        - Valid action_types: "collect arrears", "collect CPE", "collect arrears and CPE"
        - Valid statuses: "Incident Open", "Incident close", "Incident reject"
        - Date format enforcement (YYYY-MM-DD)
    - Error Handling:
        - Comprehensive validation errors
        - Database operation failures
        - File system permissions
    - Reporting:
        - Automatic filename generation with timestamp
        - Empty dataset handling
        - Console and log feedback

4. Configuration:
    - Export path determined by ConfigLoaderSingleton
    - Styles managed through style_loader.py
    - Column headers defined in INCIDENT_HEADERS constant

5. Integration Points:
    - Called by TaskHandlers.handle_task_20()
    - Uses MongoDBConnectionSingleton for database access
    - Leverages application-wide logging

Technical Specifications:
    - Input Parameters:
        - action_type: String (predefined values)
        - status: String (predefined values)
        - from_date/to_date: String (YYYY-MM-DD format)
    - Output:
        - Excel file with standardized naming convention
        - Returns boolean success status
    - Collections Accessed:
        - Incident_log (primary data source)
'''


import logging
from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from pymongo import MongoClient
from utils.connectionMongo import MongoDBConnectionSingleton
from utils.logger import SingletonLogger
from logging import getLogger
import platform
from utils.config_loader import ConfigLoaderSingleton

logger = getLogger('appLogger')

INCIDENT_HEADERS = [
    "Incident_Id", "Account_Num", "Incident_Status", "Actions",
    "Monitor_Months", "Created_By", "Created_Dtm", "Source_Type"
]

def excel_incident_detail(action_type, status, from_date, to_date):

    """Fetch and export incidents with a fixed Task_Id of 20 based on validated parameters"""

   
    try:   
            # Get export directory from config
            export_dir = ConfigLoaderSingleton().get_export_path()
            export_dir.mkdir(parents=True, exist_ok=True)

            db = MongoDBConnectionSingleton().get_database()
            incident_log_collection = db["Incident_log"]
            incident_query = {} 

            # Check each parameter and build query
            # Check action_type
            if action_type is not None:
                if action_type == "collect arrears and CPE":
                    incident_query["Actions"] = {"$regex": f"^{action_type}$"}
                elif action_type == "collect arrears":
                    incident_query["Actions"] = action_type
                elif action_type == "collect CPE":
                    incident_query["Actions"] = action_type
                else:
                    raise ValueError(f"Invalid action_type '{action_type}'. Must be 'collect arrears and CPE', 'collect arrears', or 'collect CPE'")
            

            # Check status
            if status is not None:
                if status == "Incident Open":
                    incident_query["Incident_Status"] = {"$regex": f"^{status}$"}
                elif status == "Reject":
                    incident_query["Incident_Status"] = status
                elif status == "Complete":
                    incident_query["Incident_Status"] = status
                elif status == "Incident Error":
                    incident_query["Incident_Status"] = status
                elif status == "Incident Inprogress":
                    incident_query["Incident_Status"] = status
                else:
                    raise ValueError(f"Invalid status '{status}'. Must be 'Incident Open', 'Incident Close', or 'Incident Reject'")



            # Check date range
            if from_date is not None and to_date is not None:
                try:
                    # Check if from_date and to_date are in correct YYYY-MM-DD format
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                    
                    # Validate date range
                    if to_dt < from_dt:
                        raise ValueError("to_date cannot be earlier than from_date")
                    
                   # Construct query                  
                    incident_query["Created_Dtm"] = {"$gte": from_dt, "$lte": to_dt}

                except ValueError as ve:
                    if str(ve).startswith("to_date"):
                        raise
                    raise ValueError(f"Invalid date format. Use 'YYYY-MM-DD'. Error: {str(ve)}")
            
            # Log and execute query
            logger.info(f"Executing query: {incident_query}")
            incidents = list(incident_log_collection.find(incident_query))  # Fetch data into an array
            logger.info(f"Found {len(incidents)} matching incidents")

            # Export to Excel even if no incidents are found
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
            filename = f"incidents_details_{timestamp}.xlsx"
            filepath = export_dir / filename

            wb = Workbook()
            wb.remove(wb.active)

            if not create_incident_table(wb, incidents, {
                "action": action_type,
                "status": status,
                "date_range": (from_dt if from_date is not None else None, to_dt if to_date is not None else None)
            }):
                raise Exception("Failed to create incident sheet")

            wb.save(filepath)

            # Write export record to Download collection
            try:
                download_collection = db["file_download_log"]
                export_record = {
                    "File_Name": filename,
                    "File_Path": str(filepath),
                    "Export_Timestamp": datetime.now(),
                    "Exported_Record_Count": len(incidents),
                    "Applied_Filters": {
                        "Action": action_type,
                        "Status": status,
                        "From_Date": from_date,
                        "To_Date": to_date
                    }
                }
                download_collection.insert_one(export_record)
                logger.info("Export details written to Download collection.")
            except Exception as e:
                logger.error(f"Failed to insert download record: {str(e)}", exc_info=True)


            if not incidents:
                print("No incidents found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(incidents)} records to: {filepath}")
            return True

    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}")
        print(f"Error: {str(ve)}")
        return False
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False
    

def create_incident_table(wb, data, filters=None):
    """Create formatted Excel sheet with filtered incident data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="INCIDENT REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(INCIDENT_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="INCIDENT REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            if filters.get('action'):
                ws.cell(row=row_idx, column=2, value="Action:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['action']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('status'):
                ws.cell(row=row_idx, column=2, value="Status:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['status']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('date_range') and any(filters['date_range']):
                start, end = filters['date_range']
                ws.cell(row=row_idx, column=2, value="Date Range:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                date_str = f"{start.strftime('%Y-%m-%d') if start else 'Beginning'} to {end.strftime('%Y-%m-%d') if end else 'Now'}"
                ws.cell(row=row_idx, column=3, value=date_str).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(INCIDENT_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(INCIDENT_HEADERS, 1):
                    value = record.get(header, "")
                    if header == "Incident_Id" and isinstance(value, ObjectId):
                        value = str(value)
                    if header == "Created_Dtm" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(INCIDENT_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(INCIDENT_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating sheet: {str(e)}", exc_info=True)
        return False