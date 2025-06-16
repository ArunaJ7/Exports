
from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from utils.connectionMongo import MongoDBConnectionSingleton
from logging import getLogger
from tasks.config_loader import ConfigLoaderSingleton

logger = getLogger('appLogger')

DISTRIBUTION_TRANSACTION_BATCH_HEADERS = [
    "Batch Sequence","rulebase count"
    "Approved on", 
]

def excel_case_distribution_transaction_batch_detail(case_distribution_batch_id):
    """Fetch and export case distribution DRC transaction data based on validated parameters"""
   
    try:   
            # Get export directory from config
            export_dir = ConfigLoaderSingleton().get_export_path()
            export_dir.mkdir(parents=True, exist_ok=True)

            db = MongoDBConnectionSingleton().get_database()
            case_distribution_collection = db["Case_distribution_drc_transactions"]
            drc_transaction_batch_list_query = {}

            # Check Arrears_band parameter
            if case_distribution_batch_id is not None:
                if case_distribution_batch_id == 1:
                    drc_transaction_batch_list_query["case_distribution_batch_id"] = case_distribution_batch_id
                elif case_distribution_batch_id == 2:
                    drc_transaction_batch_list_query["case_distribution_batch_id"] = case_distribution_batch_id
               
                else:
                    raise ValueError(f"Invalid case distribution batch id '{case_distribution_batch_id}'. Must be one of: 1, 2")

           

            
            # Log and execute query
            logger.info(f"Executing query: {drc_transaction_batch_list_query}")
            distributions = list(case_distribution_collection.find(drc_transaction_batch_list_query))
            logger.info(f"Found {len(distributions)} matching distributions")

            # Export to Excel even if no distributions are found
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
            filename = f"case_distribution_drc_transaction_batch_list_details_{timestamp}.xlsx"
            filepath = export_dir / filename

            wb = Workbook()
            wb.remove(wb.active)

            if not create_distribution_table(wb, distributions, {
                "case_distribution_batch_id": case_distribution_batch_id
            }):
                raise Exception("Failed to create distribution drc transaction sheet")

            wb.save(filepath)

             # Write export record to Download collection
            try:
                download_collection = db["download"]
                export_record = {
                    "File_Name": filename,
                    "File_Path": str(filepath),
                    "Export_Timestamp": datetime.now(),
                    "Exported_Record_Count": len(distributions),
                    "Applied_Filters": {
                        "Case_distribution_batch_id": case_distribution_batch_id
                    }
                }
                download_collection.insert_one(export_record)
                logger.info("Export details written to Download collection.")
            except Exception as e:
                logger.error(f"Failed to insert download record: {str(e)}", exc_info=True)


            if not distributions:
                print(f"No distributions found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(distributions)} records to: {filepath}")
            return True

    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}")
        print(f"Error: {str(ve)}")
        return False
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False
       

def create_distribution_table(wb, data, filters=None):
    """Create formatted Excel sheet with filtered distribution data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="CASE DISTRIBUTION REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DISTRIBUTION_TRANSACTION_BATCH_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="CASE DISTRIBUTION DRC TRANSACTION LIST")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            if filters.get('arrears_band'):
                ws.cell(row=row_idx, column=2, value="Arrears Band:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['arrears_band']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('drc_rule'):
                ws.cell(row=row_idx, column=2, value="DRC Commission Rule:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['drc_rule']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('date_range') and any(filters['date_range']):
                start, end = filters['date_range']
                ws.cell(row=row_idx, column=2, value="Date Range:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                date_str = f"{start.strftime('%Y-%m-%d') if start else 'Beginning'} to {end.strftime('%Y-%m-%d') if end else 'Now'}"
                ws.cell(row=row_idx, column=3, value=date_str).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(DISTRIBUTION_TRANSACTION_BATCH_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(DISTRIBUTION_TRANSACTION_BATCH_HEADERS, 1):
                    value = record.get(header, "")
                    # Handle date fields
                    if header == "Created Dtm" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    # Format case count as integer
                    if header == "Case Count" and isinstance(value, (int, float)):
                        value = int(value)
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(DISTRIBUTION_TRANSACTION_BATCH_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(DISTRIBUTION_TRANSACTION_BATCH_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating sheet: {str(e)}", exc_info=True)
        return False