'''
Purpose: This module handles the export of DRC summary data from MongoDB to formatted Excel reports
Created Date: 2025-01-18
Created By: Aruna Jayaweera (ajayaweerau@gmail.com)
Last Modified Date: 2024-02-20
Modified By: Aruna Jayaweera (ajayaweerau@gmail.com)
Version: Python 3.12
Dependencies:
    - pymongo (for MongoDB connectivity)
    - openpyxl (for Excel file operations)
    - python-dotenv (for environment variables)
    - logging (for execution tracking)
Related Files:
    - task_handler.py (initiates the export process)
    - config_loader.py (provides export path configuration)
    - style_loader.py (handles Excel styling)
    - connectionMongo.py (database connection handler)

Program Description:
1. Core Functionality:
    - excel_drc_summary_detail(): Main export function that:
        a. Validates input parameters (DRC identifier, batch ID)
        b. Constructs MongoDB query for DRC summary data
        c. Executes query against Case_Distribution_DRC_Summary collection
        d. Generates formatted Excel report
    - create_drc_summary_table(): Handles Excel sheet creation with:
        a. Professional formatting and styling
        b. Dynamic column sizing
        c. Filter headers display
        d. Empty dataset handling

2. Data Flow:
    - Receives filter parameters from calling function
    - Fetches data from Case_Distribution_DRC_Summary collection
    - Transforms MongoDB documents to Excel rows with proper formatting
    - Applies consistent styling using STYLES configuration
    - Saves report to configured export directory

3. Key Features:
    - Parameter Validation:
        - Valid DRC identifiers: "D1", "D2"
        - Valid batch IDs: 1, 2, 3
    - Data Formatting:
        - Formats timestamps (YYYY-MM-DD HH:MM:SS)
        - Converts ObjectId to string
    - Error Handling:
        - Comprehensive validation errors
        - Database operation failures
        - File system permissions
    - Reporting:
        - Automatic filename generation with timestamp (drc_summary_[timestamp].xlsx)
        - Empty dataset handling
        - Console and log feedback

4. Configuration:
    - Export path determined by ConfigLoaderSingleton
    - Styles managed through style_loader.py
    - Column headers defined in DRC_SUMMARY_HEADERS constant:
        * created_dtm
        * drc_id
        * drc
        * case_count
        * tot_arrease
        * proceed_on

5. Integration Points:
    - Called by task handlers for DRC summary reporting
    - Uses MongoDBConnectionSingleton for database access
    - Leverages application-wide logging

Technical Specifications:
    - Input Parameters:
        - drc: String (predefined values)
        - case_distribution_batch_id: Integer (predefined values)
    - Output:
        - Excel file with standardized naming convention
        - Returns boolean success status
    - Collections Accessed:
        - Case_Distribution_DRC_Summary (primary data source)
    - Query Logic:
        - Uses exact matching for DRC and batch ID values
        - Special handling for timestamp fields
'''

from datetime import datetime
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from utils.connectionMongo import MongoDBConnectionSingleton
from logging import getLogger
from utils.config_loader import ConfigLoaderSingleton

logger = getLogger('appLogger')


DRC_SUMMARY_HEADERS = [
    "created_dtm", "drc_id", "drc", "case_count", "tot_arrease", "proceed_on"
]

def excel_drc_summary_detail(drc, case_distribution_batch_id):
    """Fetch and export DRC summary details with a fixed Task_Id of 20 based on validated parameters"""
    
    
    try:
            
            # Get export directory from config
            export_dir = ConfigLoaderSingleton().get_export_path()
            export_dir.mkdir(parents=True, exist_ok=True)

            db = MongoDBConnectionSingleton().get_database()
            case_distribution_collection = db["Case_Distribution_DRC_Summary"]
            case_distribution_query = {}

            # Check each parameter and build query

            # check drc
            if drc is not None:
                if drc == "D1":
                    case_distribution_query[drc] = {"$regex": f"^{drc}$"}
                elif drc == "D2":
                    case_distribution_query[drc] = drc
                else:
                    raise ValueError(f"Invalid drc '{drc}'. Must be 'D1', or 'D2'")
            

            # check case_distribution_batch_id 
            if case_distribution_batch_id is not None:
                if case_distribution_batch_id == 1:
                    case_distribution_query[case_distribution_batch_id] = {"$regex": f"^{case_distribution_batch_id}$"}
                elif case_distribution_batch_id == 2:
                    case_distribution_query[case_distribution_batch_id] = case_distribution_batch_id
                elif case_distribution_batch_id == 3:
                    case_distribution_query[case_distribution_batch_id] = case_distribution_batch_id
                else:
                    raise ValueError(f"Invalid case distribution batch id '{case_distribution_batch_id}'. Must be 1, 2, or 3")


            #log and excute query
            logger.info(f"Executing query on Case_Distribution_DRC_Summary: {case_distribution_query}")
            summaries = list(case_distribution_collection.find(case_distribution_query)) #fetch data into array
            logger.info(f"Found {len(summaries)} matching DRC summary records")

            if not summaries:
                print("No DRC summary records found matching the selected filters")
                return False

            # Export to Excel even if no incidents are found
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"drc_summary_{timestamp}.xlsx"
            filepath = export_dir / filename

            wb = Workbook()
            wb.remove(wb.active)

            if not create_drc_summary_table(wb, summaries, {
                "drc": drc,
                "case_distribution_batch_id": case_distribution_batch_id
            }):
                raise Exception("Failed to create DRC summary sheet")

            wb.save(filepath)
            if not summaries:
                print("No drc summaries found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(summaries)} DRC summary records to: {filepath}")
            return True

    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}")
        print(f"Error: {str(ve)}")
        return False
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False
        
        

def create_drc_summary_table(wb, data, filters=None):
    """Create formatted Excel sheet with DRC summary data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="DRC SUMMARY REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DRC_SUMMARY_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="DRC SUMMARY REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            # Task ID filter
            if filters.get('task_id'):
                ws.cell(row=row_idx, column=2, value="Task ID:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=str(filters['task_id'])).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            # DRC filter
            if filters.get('drc'):
                ws.cell(row=row_idx, column=2, value="DRC:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['drc']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            # Case Distribution Batch ID filter
            if filters.get('case_distribution_batch_id') is not None:
                ws.cell(row=row_idx, column=2, value="Case Distribution Batch ID:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=str(filters['case_distribution_batch_id'])).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(DRC_SUMMARY_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(DRC_SUMMARY_HEADERS, 1):
                    value = record.get(header, "")
                    if header == "drc_id" and isinstance(value, ObjectId):
                        value = str(value)
                    if header == "created_dtm" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    if header == "proceed_on" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(DRC_SUMMARY_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(DRC_SUMMARY_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating DRC summary sheet: {str(e)}", exc_info=True)
        return False