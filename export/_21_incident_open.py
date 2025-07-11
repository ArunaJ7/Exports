import logging
from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from logging import getLogger
from pymongo import MongoClient
from utils.connectionMongo import MongoDBConnectionSingleton
from utils.config_loader import ConfigLoaderSingleton


logger = getLogger('excel_data_writer')

INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS = [
    "Id", "Incident_Status", "Account_Num", "Actions",
    "Amount", "Source_Type"
]

def excel_incident_open_distribution():
    """Fetch and export all open incidents for distribution without parameter filtering"""
    
    try:
            # Get export directory from config
            export_dir = ConfigLoaderSingleton().get_export_path()
            export_dir.mkdir(parents=True, exist_ok=True)

            db = MongoDBConnectionSingleton().get_database()
            incident_log_collection = db["Incident_log"]
            
            incident_open_query = {"Incident_Status": "Incident Open"}  # Fixed filter for open incidents

            # Log and execute query
            logger.info(f"Executing query: {incident_open_query}")
            incidents = list(incident_log_collection.find(incident_open_query))
            logger.info(f"Found {len(incidents)} matching incidents")

            # Export to Excel even if no open incidents are found
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
            filename = f"incident_open_distribution_{timestamp}.xlsx"
            filepath = export_dir / filename
         

            wb = Workbook()
            wb.remove(wb.active)

            if not create_incident_open_distribution_table(wb, incidents):
                raise Exception("Failed to create incident open distribution sheet")

            wb.save(filepath)

            # Write export record to Download collection
            try:
                download_collection = db["file_download_log"]
                export_record = {
                    "File_Name": filename,
                    "File_Path": str(filepath),
                    "Export_Timestamp": datetime.now(),
                    "Exported_Record_Count": len(incidents)
                }
                
                download_collection.insert_one(export_record)
                logger.info("Export details written to Download collection.")
            except Exception as e:
                logger.error(f"Failed to insert download record: {str(e)}", exc_info=True)


            if not incidents:
                print(f"No open incidents found. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(incidents)} records to: {filepath}")
            return True

    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False
        

def create_incident_open_distribution_table(wb, data):
    """Create formatted Excel sheet with open incident distribution data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="OPEN INCIDENT DISTRIBUTION")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="OPEN INCIDENT DISTRIBUTION REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 2
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS, 1):
                    value = record.get(header, "")
                    if header == "Id" and isinstance(value, ObjectId):
                        value = str(value)
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(INCIDENT_OPEN_FOR_DISTRIBUTION_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating sheet: {str(e)}", exc_info=True)
        return False