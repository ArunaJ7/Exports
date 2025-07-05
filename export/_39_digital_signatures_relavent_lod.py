'''
Purpose: This module handles the export of rejected incident data from MongoDB to formatted Excel reports
Created Date: 2025-03-20  
Created By: Aruna Jayaweera (ajayaweerau@gmail.com)
Last Modified Date: 2024-05-20
Modified By: Aruna Jayaweera (ajayaweerau@gmail.com)
Version: Python 3.12
Dependencies:
    - pymongo (for MongoDB connectivity)
    - openpyxl (for Excel file operations) 
    - python-dotenv (for environment variables)
    - logging (for execution tracking)
Related Files:
    - task_handler.py (initiates the export process)
    - config_loader.py (provides export path configuration)
    - style_loader.py (handles Excel styling)
    - connectionMongo.py (database connection handler)

Program Description:
1. Core Functionality:
    - excel_rejected_detail(): Main export function that:
        a. Validates input parameters (actions, DRC commission rule, date range)
        b. Constructs MongoDB query for rejected case
        c. Executes query against Incident collection
        d. Generates formatted Excel report
    - create_rejected_table(): Handles Excel sheet creation with:
        a. Professional formatting and styling
        b. Dynamic column sizing
        c. Filter headers display
        d. Empty dataset handling

2. Data Flow:
    - Receives filter parameters from calling function
    - Fetches data from Incident collection with "Incident_Status": "Incident Reject"
    - Transforms MongoDB documents to Excel rows
    - Applies consistent styling using STYLES configuration
    - Saves report to configured export directory

3. Key Features:
    - Parameter Validation:
        - Valid actions: "collect arrears and CPE", "collect arrears", "collect CPE"
        - Valid DRC commission rules: "PEO TV" or "BB"
        - Date format enforcement (YYYY-MM-DD)
        - Date range validation (to_date cannot be earlier than from_date)
    - Error Handling:
        - Comprehensive validation errors
        - Database operation failures
        - File system permissions
    - Reporting:
        - Automatic filename generation with timestamp (rejected_incidents_[timestamp].xlsx)
        - Empty dataset handling
        - Console and log feedback

4. Configuration:
    - Export path determined by ConfigLoaderSingleton
    - Styles managed through style_loader.py
    - Column headers defined in DIGITAL_SIGNATURES_HEADERS)) constant:
        * Incident_Id
        * Incident_Status
        * Account_Num
        * Created_Dtm
        * Filtered_Reason
        * Rejected_Dtm
        * Rejected_By

5. Integration Points:
    - Called by task handlers for rejected incident reporting
    - Uses MongoDBConnectionSingleton for database access
    - Leverages application-wide logging

Technical Specifications:
    - Input Parameters:
        - actions: String (predefined values)
        - drc_commision_rule: String ("PEO TV" or "BB")
        - from_date/to_date: String (YYYY-MM-DD format)
    - Output:
        - Excel file with standardized naming convention
        - Returns boolean success status
    - Collections Accessed:
        - Incident (primary data source)
        - Filters specifically for "Incident_Status": "Incident Reject"
    - Special Data Handling:
        - Converts ObjectId to string for Incident_Id
        - Formats datetime objects for Created_Dtm and Rejected_Dtm
'''

from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from utils.connectionMongo import MongoDBConnectionSingleton
from logging import getLogger
from utils.config_loader import ConfigLoaderSingleton

logger = getLogger('appLogger')

DIGITAL_SIGNATURES_HEADERS = [
    "Incident_Id", "Incident_Status", "Account_Num", "Created_Dtm",
    "Filtered_Reason"]

def excel_digital_signature_detail(case_current_status):
    """Fetch and export digital signature details from Incident collection"""

    try:
            # Get export directory from config
            export_dir = ConfigLoaderSingleton().get_export_path()
            export_dir.mkdir(parents=True, exist_ok=True)

            db = MongoDBConnectionSingleton().get_database()
            case_details_collection = db["case_details"]
            case_current_query = {}  

            # Validate and apply actions filter
            if case_current_status is not None:
                if case_current_status == "Abandand":
                    case_current_query["Case_current_starus"] = {"$regex": f"^{case_current_status}$"}
                elif case_current_status == "LIT prescribed":
                    case_current_query["Case_current_starus"] = case_current_status
                else:
                     raise ValueError(f"Invalid actions '{case_current_status}'. Must be 'Abandand', 'LIT prescribed'")

            

           


            logger.info(f"Executing query on Incident for rejected case: {case_current_query}")
            case = list(case_details_collection.find(case_current_query))
            logger.info(f"Found {len(case)} matching rejected case")

            # Export to Excel
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
            filename = f"digital_signatures_relavent_lod_{timestamp}.xlsx"
            filepath = export_dir / filename

            wb = Workbook()
            wb.remove(wb.active)

            if not create_digital_signature_table(wb, case, {
                "Case_current_status": case_current_status

            }):
                raise Exception("Failed to create digital signatures sheet")

            wb.save(filepath)

             # Write export record to Download collection
            try:
                download_collection = db["file_download_log"]
                export_record = {
                    "File_Name": filename,
                    "File_Path": str(filepath),
                    "Export_Timestamp": datetime.now(),
                    "Exported_Record_Count": len(case),
                    "Applied_Filters": {
                        "Case_current_status" : case_current_status
                    }
                }
                download_collection.insert_one(export_record)
                logger.info("Export details written to Download collection.")
            except Exception as e:
                logger.error(f"Failed to insert download record: {str(e)}", exc_info=True)


            if not case:
                print("No digital signatures found matching the selected filters. Exported empty table to: {filepath}")
            else:    
                print(f"\nSuccessfully exported {len(case)} signatures records to: {filepath}")
            return True            
           
    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}")
        print(f"Error: {str(ve)}")
        return False
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False
    
        

def create_digital_signature_table(wb, data, filters=None):
    """Create formatted Excel sheet with digital signature data"""
    try:
        ws = wb.create_sheet(title="DIGITAL SIGNATURES RELAVENT LOD REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DIGITAL_SIGNATURES_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="DIGITAL SIGNATURES RELAVENT LOD REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            # Case_current_starus filter
            if filters.get('actions'):
                ws.cell(row=row_idx, column=2, value="Case_current_status:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['Case_current_status']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(DIGITAL_SIGNATURES_HEADERS)), 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows
        for record in data:
            row_idx += 1
            for col_idx, header in enumerate(DIGITAL_SIGNATURES_HEADERS)), 1):
                value = record.get(header, "")
                if header == "Incident_Id" and isinstance(value, ObjectId):
                    value = str(value)
                if header == "Created_Dtm" and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = STYLES['Border_Style']['font']
                cell.border = STYLES['Border_Style']['border']
                cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to all columns
        if data:
            last_col_letter = get_column_letter(len(DIGITAL_SIGNATURES_HEADERS))
            ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{row_idx}"
        
        # Auto-adjust columns
        for col_idx in range(1, len(DIGITAL_SIGNATURES_HEADERS)) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating digital signature sheet: {str(e)}", exc_info=True)
        return False    