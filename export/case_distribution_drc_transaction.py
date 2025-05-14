import logging
from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
import os
from utils.connectDB import get_db_connection
import logging.config
from utils.config_loader import get_config
from pymongo import MongoClient

logger = logging.getLogger('excel_data_writer')

DISTRIBUTION_HEADERS = [
    "Case Distribution Batch ID", "Created Dtm", "Distributed Status",
    "Action Type", "DRC Commission Rule", "Arrears Band", 
    "Case Count", "Approval"
]

def excel_case_distribution_detail(Arrears_band, drc_commision_rule, date_from, date_to):
    """Fetch and export case distribution DRC transaction data based on validated parameters"""
    try:
        client = MongoClient("mongodb://localhost:27017/")
        db = client["DRS"]
        logger.info(f"Connected to MongoDB successfully | DRS")

    except Exception as err:
        print("Connection error")
        logger.error(f"MongoDB connection failed: {str(err)}")
        return False
    else:
        try:   
            collection = db["Case_Distribution_log"]
            query = {}

            # Check Arrears_band parameter
            if Arrears_band is not None:
                valid_arrears_bands = ["0-30", "31-60", "61-90", "91+"]  # Example bands, adjust as needed
                if Arrears_band not in valid_arrears_bands:
                    raise ValueError(f"Invalid Arrears_band '{Arrears_band}'. Must be one of: {', '.join(valid_arrears_bands)}")
                query["Arrears Band"] = Arrears_band

            # Check drc_commision_rule parameter
            if drc_commision_rule is not None:
                if not isinstance(drc_commision_rule, str) or not drc_commision_rule.strip():
                    raise ValueError("drc_commision_rule must be a non-empty string")
                query["DRC Commission Rule"] = {"$regex": f"^{drc_commision_rule.strip()}$", "$options": "i"}

            # Check date range
            if date_from is not None and date_to is not None:
                try:
                    # Check if dates are in correct YYYY-MM-DD format
                    from_dt = datetime.strptime(date_from, '%Y-%m-%d')
                    to_dt = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                    
                    # Validate date range
                    if to_dt < from_dt:
                        raise ValueError("date_to cannot be earlier than date_from")
                    
                    # Construct query for date range (checking Created Dtm)
                    query["Created Dtm"] = {"$gte": from_dt, "$lte": to_dt}

                except ValueError as ve:
                    if str(ve).startswith("date_to"):
                        raise
                    raise ValueError(f"Invalid date format. Use 'YYYY-MM-DD'. Error: {str(ve)}")

            # Log and execute query
            logger.info(f"Executing query: {query}")
            distributions = list(collection.find(query))
            logger.info(f"Found {len(distributions)} matching distributions")

            # Export to Excel even if no distributions are found
            output_dir = "exports"
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"case_distribution_details_{timestamp}.xlsx"
            filepath = os.path.join(output_dir, filename)
            os.makedirs(output_dir, exist_ok=True)

            wb = Workbook()
            wb.remove(wb.active)

            if not create_distribution_table(wb, distributions, {
                "arrears_band": Arrears_band,
                "drc_rule": drc_commision_rule,
                "date_range": (from_dt if date_from is not None else None, to_dt if date_to is not None else None)
            }):
                raise Exception("Failed to create distribution sheet")

            wb.save(filepath)
            if not distributions:
                print(f"No distributions found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(distributions)} records to: {filepath}")
            return True

        except ValueError as ve:
            logger.error(f"Validation error: {str(ve)}")
            print(f"Error: {str(ve)}")
            return False
        except Exception as e:
            logger.error(f"Export failed: {str(e)}", exc_info=True)
            print(f"\nError during export: {str(e)}")
            return False
        finally:
            if client:
                client.close()
                logger.info("MongoDB connection closed")

def create_distribution_table(wb, data, filters=None):
    """Create formatted Excel sheet with filtered distribution data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="CASE DISTRIBUTION REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DISTRIBUTION_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="CASE DISTRIBUTION DRC TRANSACTION LIST")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            if filters.get('arrears_band'):
                ws.cell(row=row_idx, column=2, value="Arrears Band:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['arrears_band']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('drc_rule'):
                ws.cell(row=row_idx, column=2, value="DRC Commission Rule:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['drc_rule']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('date_range') and any(filters['date_range']):
                start, end = filters['date_range']
                ws.cell(row=row_idx, column=2, value="Date Range:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                date_str = f"{start.strftime('%Y-%m-%d') if start else 'Beginning'} to {end.strftime('%Y-%m-%d') if end else 'Now'}"
                ws.cell(row=row_idx, column=3, value=date_str).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(DISTRIBUTION_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(DISTRIBUTION_HEADERS, 1):
                    value = record.get(header, "")
                    # Handle date fields
                    if header == "Created Dtm" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    # Format case count as integer
                    if header == "Case Count" and isinstance(value, (int, float)):
                        value = int(value)
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(DISTRIBUTION_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(DISTRIBUTION_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating sheet: {str(e)}", exc_info=True)
        return False