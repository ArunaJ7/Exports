'''
Purpose: This module handles the export of case distribution transaction data from MongoDB to formatted Excel reports
Created Date: 2025-01-18  
Created By: Aruna Jayaweera (ajayaweerau@gmail.com)
Last Modified Date: 2024-02-20  
Modified By: Aruna Jayaweera (ajayaweerau@gmail.com)
Version: Python 3.12

Dependencies:
    - pymongo (for MongoDB connectivity)
    - openpyxl (for Excel file operations)
    - python-dotenv (for environment variables)
    - logging (for execution tracking)

Related Files:
    - task_handler.py (initiates the export process)
    - config_loader.py (provides export path configuration)
    - style_loader.py (handles Excel styling)
    - connectionMongo.py (database connection handler)

Program Description:
1. Core Functionality:
    - excel_case_distribution_detail(): Main export function that:
        a. Validates input parameters (arrears band, DRC rule, date range)
        b. Constructs MongoDB query for distribution data
        c. Executes query against Case_Distribution_log collection
        d. Generates formatted Excel report
    - create_distribution_table(): Handles Excel sheet creation with:
        a. Professional formatting and styling
        b. Dynamic column sizing
        c. Filter headers display
        d. Empty dataset handling

2. Data Flow:
    - Receives filter parameters from calling function
    - Fetches data from Case_Distribution_log collection
    - Transforms MongoDB documents to Excel rows with proper formatting
    - Applies consistent styling using STYLES configuration
    - Saves report to configured export directory

3. Key Features:
    - Parameter Validation:
        - Valid arrears bands: "0-30", "31-60", "61-90", "91+"
        - Validates DRC commission rule as non-empty string
        - Date format enforcement (YYYY-MM-DD)
        - Date range validation (date_to cannot be earlier than date_from)
    - Data Formatting:
        - Formats datetime objects for Created Dtm
        - Converts Case Count to integers
    - Error Handling:
        - Comprehensive validation errors
        - Database operation failures
        - File system permissions
    - Reporting:
        - Automatic filename generation with timestamp (case_distribution_details_[timestamp].xlsx)
        - Empty dataset handling with headers
        - Console and log feedback

4. Configuration:
    - Export path determined by ConfigLoaderSingleton
    - Styles managed through style_loader.py
    - Column headers defined in DISTRIBUTION_HEADERS constant:
        * Case Distribution Batch ID
        * Created Dtm
        * Distributed Status
        * Action Type
        * DRC Commission Rule
        * Arrears Band
        * Case Count
        * Approval

5. Integration Points:
    - Called by task handlers for case distribution reporting
    - Uses MongoDBConnectionSingleton for database access
    - Leverages application-wide logging

Technical Specifications:
    - Input Parameters:
        - Arrears_band: String (predefined values)
        - drc_commision_rule: String
        - date_from/date_to: String (YYYY-MM-DD format)
    - Output:
        - Excel file with standardized naming convention
        - Returns boolean success status
    - Collections Accessed:
        - Case_Distribution_log (primary data source)
    - Query Logic:
        - Case-sensitive regex matching for DRC rules
        - Exact matching for arrears bands
        - Date range filtering on Created Dtm field
'''

from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from utils.connectionMongo import MongoDBConnectionSingleton
from logging import getLogger
from utils.config_loader import ConfigLoaderSingleton

logger = getLogger('appLogger')

DISTRIBUTION_HEADERS = [
    "Case Distribution Batch ID", "Created Dtm", "Distributed Status",
    "Action Type", "DRC Commission Rule", "Arrears Band", 
    "Case Count", "Approval"
]

def excel_case_distribution_detail(arrears_band, drc_commision_rule, from_date, to_date):
    """Fetch and export case distribution DRC transaction data based on validated parameters"""
   
    try:   
            # Get export directory from config
            export_dir = ConfigLoaderSingleton().get_export_path()
            export_dir.mkdir(parents=True, exist_ok=True)

            db = MongoDBConnectionSingleton().get_database()
            case_distribution_collection = db["Case_distribution_drc_transactions"]
            drc_transaction_query = {}

            # Check Arrears_band parameter
            if arrears_band is not None:
                if arrears_band == "AB-5_10":
                    drc_transaction_query["Arrears Band"] = arrears_band
                elif arrears_band == "AB-25_50":
                    drc_transaction_query["Arrears Band"] = arrears_band
                else:
                    raise ValueError(f"Invalid Arrears_band '{arrears_band}'. Must be one of: AB-5_10, AB-25_50")

            # Check drc_commision_rule parameter
            if drc_commision_rule is not None:
                if drc_commision_rule == "PEO TV":
                    drc_transaction_query["drc_commision_rule"] = {"$regex": f"^{drc_commision_rule}$"}
                elif drc_commision_rule == "BB":
                    drc_transaction_query["drc_commision_rule"] = drc_commision_rule
                else:
                    raise ValueError(f"Invalid drc_commision_rule '{drc_commision_rule}'. Must be 'PEO TV', 'BB'")
            

            # Check date range
            if from_date is not None and to_date is not None:
                try:
                    # Check if dates are in correct YYYY-MM-DD format
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                    
                    # Validate date range
                    if to_dt < from_dt:
                        raise ValueError("to_date cannot be earlier than from_date")
                    
                    # Construct query for date range (checking Created Dtm)
                    drc_transaction_query["Created Dtm"] = {"$gte": from_dt, "$lte": to_dt}

                except ValueError as ve:
                    if str(ve).startswith("to_date"):
                        raise
                    raise ValueError(f"Invalid date format. Use 'YYYY-MM-DD'. Error: {str(ve)}")

            # Log and execute query
            logger.info(f"Executing query: {drc_transaction_query}")
            distributions = list(case_distribution_collection.find(drc_transaction_query))
            logger.info(f"Found {len(distributions)} matching distributions")

            # Export to Excel even if no distributions are found
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
            filename = f"case_distribution_details_{timestamp}.xlsx"
            filepath = export_dir / filename

            wb = Workbook()
            wb.remove(wb.active)

            if not create_distribution_table(wb, distributions, {
                "arrears_band": arrears_band,
                "drc_rule": drc_commision_rule,
                "date_range": (from_dt if from_date is not None else None, to_dt if to_date is not None else None)
            }):
                raise Exception("Failed to create distribution sheet")

            wb.save(filepath)

             # Write export record to Download collection
            try:
                download_collection = db["file_download_log"]
                export_record = {
                    "File_Name": filename,
                    "File_Path": str(filepath),
                    "Export_Timestamp": datetime.now(),
                    "Exported_Record_Count": len(distributions),
                    "Applied_Filters": {
                        "Arrears_Band": arrears_band,
                        "DRC_commision_rule": drc_commision_rule,
                        "From_Date": from_date,
                        "To_Date": to_date
                    }
                }
                download_collection.insert_one(export_record)
                logger.info("Export details written to Download collection.")
            except Exception as e:
                logger.error(f"Failed to insert download record: {str(e)}", exc_info=True)


            if not distributions:
                print(f"No distributions found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(distributions)} records to: {filepath}")
            return True

    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}")
        print(f"Error: {str(ve)}")
        return False
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False
       

def create_distribution_table(wb, data, filters=None):
    """Create formatted Excel sheet with filtered distribution data, including headers even if no data"""
    try:
        ws = wb.create_sheet(title="CASE DISTRIBUTION REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(DISTRIBUTION_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="CASE DISTRIBUTION DRC TRANSACTION LIST")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            if filters.get('arrears_band'):
                ws.cell(row=row_idx, column=2, value="Arrears Band:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['arrears_band']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('drc_rule'):
                ws.cell(row=row_idx, column=2, value="DRC Commission Rule:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['drc_rule']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            if filters.get('date_range') and any(filters['date_range']):
                start, end = filters['date_range']
                ws.cell(row=row_idx, column=2, value="Date Range:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                date_str = f"{start.strftime('%Y-%m-%d') if start else 'Beginning'} to {end.strftime('%Y-%m-%d') if end else 'Now'}"
                ws.cell(row=row_idx, column=3, value=date_str).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(DISTRIBUTION_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows (only if data exists)
        if data:
            for record in data:
                row_idx += 1
                for col_idx, header in enumerate(DISTRIBUTION_HEADERS, 1):
                    value = record.get(header, "")
                    # Handle date fields
                    if header == "Created Dtm" and isinstance(value, datetime):
                        value = value.strftime('%Y-%m-%d %H:%M:%S')
                    # Format case count as integer
                    if header == "Case Count" and isinstance(value, (int, float)):
                        value = int(value)
                    
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.font = STYLES['Border_Style']['font']
                    cell.border = STYLES['Border_Style']['border']
                    cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to headers
        last_col_letter = get_column_letter(len(DISTRIBUTION_HEADERS))
        ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{header_row}"
        
        # Auto-adjust columns based on headers (and data if present)
        for col_idx in range(1, len(DISTRIBUTION_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = max(adjusted_width, 20)
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating sheet: {str(e)}", exc_info=True)
        return False