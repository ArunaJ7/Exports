'''
Purpose: This module handles the export of CPE (Customer Premises Equipment) collection incidents from MongoDB to formatted Excel reports
Created Date: 2025-01-18
Created By: Aruna Jayaweera (ajayaweerau@gmail.com)
Last Modified Date: 2024-02-20
Modified By: Aruna Jayaweera (ajayaweerau@gmail.com)
Version: Python 3.12
Dependencies:
- pymongo (for MongoDB connectivity)
- openpyxl (for Excel file operations)
- python-dotenv (for environment variables)
- logging (for execution tracking)
Related Files:
- task_handler.py (initiates the export process)
- config_loader.py (provides export path configuration)
- style_loader.py (handles Excel styling)
- connectionMongo.py (database connection handler)

Program Description:

Core Functionality:

excel_cpe_detail(): Main export function that:
a. Validates input parameters (date range, DRC commission rule)
b. Constructs MongoDB query for CPE collection incidents
c. Executes query against Incident_log collection
d. Generates formatted Excel report

create_cpe_table(): Handles Excel sheet creation with:
a. Professional formatting and styling
b. Dynamic column sizing
c. Filter headers display
d. Empty dataset handling

Data Flow:

Receives filter parameters from calling function

Fetches data from Incident_log collection with "Actions": "collect CPE"

Transforms MongoDB documents to Excel rows

Applies consistent styling using STYLES configuration

Saves report to configured export directory

Key Features:

Parameter Validation:

Valid DRC commission rules: "PEO TV" or "BB"

Date format enforcement (YYYY-MM-DD)

Date range validation (to_date cannot be earlier than from_date)

Error Handling:

Comprehensive validation errors

Database operation failures

File system permissions

Reporting:

Automatic filename generation with timestamp (cpe_incidents_[timestamp].xlsx)

Empty dataset handling

Console and log feedback

Configuration:

Export path determined by ConfigLoaderSingleton

Styles managed through style_loader.py

Column headers defined in CPE_HEADERS constant:

Incident_Id

Incident_Status

Account_Num

Actions

Created_Dtm

Integration Points:

Called by task handlers for CPE collection reporting

Uses MongoDBConnectionSingleton for database access

Leverages application-wide logging

Technical Specifications:
- Input Parameters:
- from_date/to_date: String (YYYY-MM-DD format)
- drc_commision_rule: String ("PEO TV" or "BB")
- Output:
- Excel file with standardized naming convention
- Returns boolean success status
- Collections Accessed:
- Incident_log (primary data source)
- Filters specifically for "Actions": "collect CPE"
- Special Data Handling:
- Converts ObjectId to string for Incident_Id
- Formats datetime objects for Created_Dtm
'''

from datetime import datetime, timedelta
from bson import ObjectId
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from utils.style_loader import STYLES
from utils.connectionMongo import MongoDBConnectionSingleton
from logging import getLogger
from utils.config_loader import ConfigLoaderSingleton

logger = getLogger('appLogger')

CPE_HEADERS = [
    "Incident_Id", "Incident_Status", "Account_Num", "Actions",
    "Created_Dtm"
]

def excel_cpe_detail(from_date, to_date, drc_commision_rule):
    """Fetch and export 'collect CPE' incidents from Incident collection"""


    try:
            # Get export directory from config
            export_dir = ConfigLoaderSingleton().get_export_path()
            export_dir.mkdir(parents=True, exist_ok=True)

            db = MongoDBConnectionSingleton().get_database()
            incident_log_collection = db["Incident_log"]
            cpe_list_query = {"Actions": "collect CPE"}  # Fixed to only collect CPE

            

            # Apply date range filter
            if from_date is not None and to_date is not None:
                try:
                    # Check if from_date and to_date are in correct YYYY-MM-DD format
                    from_dt = datetime.strptime(from_date, '%Y-%m-%d')
                    to_dt = datetime.strptime(to_date, '%Y-%m-%d') + timedelta(days=1) - timedelta(seconds=1)
                    
                    # Validate date range
                    if to_dt < from_dt:
                        raise ValueError("to_date cannot be earlier than from_date")
                    
                   # Construct query                  
                    cpe_list_query["Created_Dtm"] = {"$gte": from_dt, "$lte": to_dt}

                except ValueError as ve:
                    if str(ve).startswith("to_date"):
                        raise
                    raise ValueError(f"Invalid date format. Use 'YYYY-MM-DD'. Error: {str(ve)}")

            # Validate and apply drc_commision_rule filter
            if drc_commision_rule is not None:
                if drc_commision_rule == "PEO TV":
                    cpe_list_query["Drc commision rule"] = {"$regex": f"^{drc_commision_rule}$"}
                elif drc_commision_rule == "BB":
                    cpe_list_query["Actions"] = drc_commision_rule   
                else:
                    raise ValueError(f"Invalid drc_commision_rule '{drc_commision_rule}'. Must be 'PEO TV', 'BB'")
                         

            logger.info(f"Executing query on Incident for CPE: {cpe_list_query}")
            incidents = list(incident_log_collection.find(cpe_list_query))
            logger.info(f"Found {len(incidents)} matching CPE incidents")

            # Export to Excel
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S%f")
            filename = f"cpe_incidents_{timestamp}.xlsx"
            filepath = export_dir / filename

            wb = Workbook()
            wb.remove(wb.active)

            if not create_cpe_table(wb, incidents, {
                "action": "collect CPE",
                "drc_commision_rule": drc_commision_rule,
                "date_range": (datetime.strptime(from_date, '%Y-%m-%d') if from_date else None,
                            datetime.strptime(to_date, '%Y-%m-%d') if to_date else None)
            }):
                raise Exception("Failed to create CPE incident sheet")

            wb.save(filepath)

            # Write export record to Download collection
            try:
                download_collection = db["download"]
                export_record = {
                    "File_Name": filename,
                    "File_Path": str(filepath),
                    "Export_Timestamp": datetime.now(),
                    "Exported_Record_Count": len(incidents),
                    "Applied_Filters": {
                        "From_Date": from_date,
                        "To_Date": to_date,
                        "DRC_Commsion_Rule": drc_commision_rule
                    }
                }
                download_collection.insert_one(export_record)
                logger.info("Export details written to Download collection.")
            except Exception as e:
                logger.error(f"Failed to insert download record: {str(e)}", exc_info=True)


            if not incidents:
                print("No CPE incidents found matching the selected filters. Exported empty table to: {filepath}")
            else:
                print(f"\nSuccessfully exported {len(incidents)} CPE records to: {filepath}")
            return True

    except ValueError as ve:
        logger.error(f"Validation error: {str(ve)}")
        print(f"Error: {str(ve)}")
        return False
    except Exception as e:
        logger.error(f"Export failed: {str(e)}", exc_info=True)
        print(f"\nError during export: {str(e)}")
        return False   
    

def create_cpe_table(wb, data, filters=None):
    """Create formatted Excel sheet with CPE incident data"""
    try:
        ws = wb.create_sheet(title="CPE INCIDENT REPORT")
        row_idx = 1
        
        # Main Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=len(CPE_HEADERS))
        main_header = ws.cell(row=row_idx, column=1, value="CPE INCIDENT REPORT")
        main_header.font = STYLES['MainHeader_Style']['font']
        main_header.fill = STYLES['MainHeader_Style']['fill']
        main_header.alignment = STYLES['MainHeader_Style']['alignment']
        row_idx += 1
        
        # Display Active Filters
        if filters:
            row_idx += 1
            
            # Action filter (always "collect CPE")
            ws.cell(row=row_idx, column=2, value="Action:").font = STYLES['FilterParam_Style']['font']
            ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
            ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
            ws.cell(row=row_idx, column=3, value=filters['action']).font = STYLES['FilterValue_Style']['font']
            ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
            ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
            row_idx += 1
            
            # DRC Commission Rule filter
            if filters.get('drc_commision_rule'):
                ws.cell(row=row_idx, column=2, value="DRC Commission Rule:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                ws.cell(row=row_idx, column=3, value=filters['drc_commision_rule']).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            # Date Range filter
            if filters.get('date_range') and any(filters['date_range']):
                start, end = filters['date_range']
                ws.cell(row=row_idx, column=2, value="Date Range:").font = STYLES['FilterParam_Style']['font']
                ws.cell(row=row_idx, column=2).fill = STYLES['FilterParam_Style']['fill']
                ws.cell(row=row_idx, column=2).alignment = STYLES['FilterParam_Style']['alignment']
                date_str = f"{start.strftime('%Y-%m-%d') if start else 'Beginning'} to {end.strftime('%Y-%m-%d') if end else 'Now'}"
                ws.cell(row=row_idx, column=3, value=date_str).font = STYLES['FilterValue_Style']['font']
                ws.cell(row=row_idx, column=3).fill = STYLES['FilterValue_Style']['fill']
                ws.cell(row=row_idx, column=3).alignment = STYLES['FilterValue_Style']['alignment']
                row_idx += 1
            
            row_idx += 1
        
        # Data Table Headers
        header_row = row_idx
        for col_idx, header in enumerate(CPE_HEADERS, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header.replace('_', ' ').title())
            cell.font = STYLES['SubHeader_Style']['font']
            cell.fill = STYLES['SubHeader_Style']['fill']
            cell.border = STYLES['SubHeader_Style']['border']
            cell.alignment = STYLES['SubHeader_Style']['alignment']
            ws.column_dimensions[get_column_letter(col_idx)].width = 20
        
        # Data Rows
        for record in data:
            row_idx += 1
            for col_idx, header in enumerate(CPE_HEADERS, 1):
                value = record.get(header, "")
                if header == "Incident_Id" and isinstance(value, ObjectId):
                    value = str(value)
                if header == "Created_Dtm" and isinstance(value, datetime):
                    value = value.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.font = STYLES['Border_Style']['font']
                cell.border = STYLES['Border_Style']['border']
                cell.alignment = STYLES['Border_Style']['alignment']
        
        # Add AutoFilter to all columns
        if data:
            last_col_letter = get_column_letter(len(CPE_HEADERS))
            ws.auto_filter.ref = f"{get_column_letter(1)}{header_row}:{last_col_letter}{row_idx}"
        
        # Auto-adjust columns
        for col_idx in range(1, len(CPE_HEADERS) + 1):
            col_letter = get_column_letter(col_idx)
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in ws[col_letter]
            )
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[col_letter].width = adjusted_width
        
        return True
    
    except Exception as e:
        logger.error(f"Error creating CPE sheet: {str(e)}", exc_info=True)
        return False